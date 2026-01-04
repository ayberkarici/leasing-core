from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse, HttpResponse, FileResponse, StreamingHttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.core.files.base import ContentFile
from django.db import models
from django.contrib.auth import get_user_model
from django import forms
from django.core.cache import cache
from datetime import datetime
import json
import os
import time

from core.decorators import admin_required
from core.mixins import AdminRequiredMixin

from .models import ADLogAnalysis, ADLogSourcePath, ProcessedADFile, SystemGID, GIDDiscrepancy, ADLogEmailTemplate, UsageType, BulkUserImport
from .services.ad_log_service import ADLogService
from .services.bulk_user_service import process_bulk_import
from accounts.models import Department
from customers.models import Customer

User = get_user_model()


# ===============================
# İş Kolları (Usage Types)
# ===============================

class UsageTypeListView(AdminRequiredMixin, ListView):
    """İş Kolları listesi"""
    model = UsageType
    template_name = 'it_tools/usage_type_list.html'
    context_object_name = 'usage_types'
    paginate_by = 20
    
    def get_queryset(self):
        return UsageType.objects.all().order_by('name')


class UsageTypeCreateView(AdminRequiredMixin, CreateView):
    """Yeni İş Kolu oluştur"""
    model = UsageType
    template_name = 'it_tools/usage_type_form.html'
    fields = ['name', 'code', 'description', 'is_active']
    success_url = reverse_lazy('it_tools:usage_type_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'İş Kolu başarıyla oluşturuldu.')
        return super().form_valid(form)


class UsageTypeUpdateView(AdminRequiredMixin, UpdateView):
    """İş Kolu düzenle"""
    model = UsageType
    template_name = 'it_tools/usage_type_form.html'
    fields = ['name', 'code', 'description', 'is_active']
    success_url = reverse_lazy('it_tools:usage_type_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'İş Kolu başarıyla güncellendi.')
        return super().form_valid(form)


class UsageTypeDeleteView(AdminRequiredMixin, DeleteView):
    """İş Kolu sil"""
    model = UsageType
    template_name = 'it_tools/usage_type_confirm_delete.html'
    success_url = reverse_lazy('it_tools:usage_type_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'İş Kolu başarıyla silindi.')
        return super().delete(request, *args, **kwargs)


class ITToolsDashboardView(AdminRequiredMixin, TemplateView):
    """IT Araçları ana sayfası"""
    template_name = 'it_tools/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['recent_analyses'] = ADLogAnalysis.objects.all()[:5]
        context['total_company_users'] = User.objects.exclude(user_type='customer').count()
        context['total_customers'] = Customer.objects.count()
        context['total_analyses'] = ADLogAnalysis.objects.count()
        context['pending_emails'] = ADLogAnalysis.objects.filter(status='email_pending').count()
        return context


# ===============================
# AD Log Analizleri
# ===============================

class ADLogListView(AdminRequiredMixin, ListView):
    """AD Log analiz listesi"""
    model = ADLogAnalysis
    template_name = 'it_tools/ad_log_list.html'
    context_object_name = 'analyses'
    paginate_by = 20
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import ADLogSourcePath
        context['source_paths'] = ADLogSourcePath.objects.filter(is_active=True)
        return context


class ADLogSourcePathListView(AdminRequiredMixin, ListView):
    """Path tanımları listesi - tüm tipler"""
    model = ADLogSourcePath
    template_name = 'it_tools/path_definition_list.html'
    context_object_name = 'path_definitions'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = super().get_queryset()
        usage_type = self.request.GET.get('type')
        if usage_type:
            queryset = queryset.filter(usage_type=usage_type)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import UsageType
        context['usage_types'] = UsageType.objects.filter(is_active=True)
        context['selected_type'] = self.request.GET.get('type', '')
        return context


class ADLogSourcePathCreateView(AdminRequiredMixin, CreateView):
    """Yeni path tanımı ekle"""
    model = ADLogSourcePath
    template_name = 'it_tools/path_definition_form.html'
    fields = ['name', 'usage_type', 'source_path', 'output_path', 'is_active', 'is_default']
    success_url = reverse_lazy('it_tools:source_path_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import UsageType
        context['usage_types'] = UsageType.objects.filter(is_active=True).order_by('name')
        return context
    
    def form_valid(self, form):
        messages.success(self.request, 'Path tanımı eklendi.')
        return super().form_valid(form)


class ADLogSourcePathUpdateView(AdminRequiredMixin, UpdateView):
    """Path tanımı düzenle"""
    model = ADLogSourcePath
    template_name = 'it_tools/path_definition_form.html'
    fields = ['name', 'usage_type', 'source_path', 'output_path', 'is_active', 'is_default']
    success_url = reverse_lazy('it_tools:source_path_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import UsageType
        context['usage_types'] = UsageType.objects.filter(is_active=True).order_by('name')
        return context
    
    def form_valid(self, form):
        messages.success(self.request, 'Path tanımı güncellendi.')
        return super().form_valid(form)


class ADLogSourcePathDeleteView(AdminRequiredMixin, DeleteView):
    """Path tanımı sil"""
    model = ADLogSourcePath
    template_name = 'it_tools/path_definition_confirm_delete.html'
    success_url = reverse_lazy('it_tools:source_path_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Path tanımı silindi.')
        return super().delete(request, *args, **kwargs)


class ADLogCreateView(AdminRequiredMixin, CreateView):
    """Yeni AD Log analizi oluştur"""
    model = ADLogAnalysis
    template_name = 'it_tools/ad_log_form.html'
    fields = ['name', 'description', 'source_path_config', 'year', 'month']
    success_url = reverse_lazy('it_tools:ad_log_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import ADLogSourcePath, UsageType
        # AD_LOG iş kolunu bul
        ad_log_type = UsageType.objects.filter(code='AD_LOG').first()
        if ad_log_type:
            context['source_paths'] = ADLogSourcePath.objects.filter(
                is_active=True, 
                usage_type=ad_log_type
            )
        else:
            context['source_paths'] = ADLogSourcePath.objects.none()
        # Yıl seçenekleri
        current_year = datetime.now().year
        context['year_choices'] = list(range(current_year - 2, current_year + 1))
        return context
    
    def get_initial(self):
        initial = super().get_initial()
        from .models import ADLogSourcePath, UsageType
        # Varsayılan path (AD_LOG tipi)
        ad_log_type = UsageType.objects.filter(code='AD_LOG').first()
        if ad_log_type:
            default_path = ADLogSourcePath.objects.filter(
                is_default=True, 
                usage_type=ad_log_type
            ).first()
            if default_path:
                initial['source_path_config'] = default_path
        # Varsayılan yıl ve ay
        now = datetime.now()
        initial['year'] = now.year
        initial['month'] = now.month - 1 if now.month > 1 else 12
        return initial
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'AD Log analizi oluşturuldu.')
        return super().form_valid(form)


class ADLogDetailView(AdminRequiredMixin, DetailView):
    """AD Log analiz detayı"""
    model = ADLogAnalysis
    template_name = 'it_tools/ad_log_detail.html'
    context_object_name = 'analysis'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Discrepancy'leri GID'ye göre grupla
        from collections import defaultdict
        
        all_discrepancies = self.object.discrepancies.all()
        
        # GID'lere göre grupla
        grouped_missing_in_system = defaultdict(list)
        missing_in_ad = []
        
        for d in all_discrepancies:
            if d.discrepancy_type == 'missing_in_system':
                grouped_missing_in_system[d.gid].append(d.source_file)
            elif d.discrepancy_type == 'missing_in_ad':
                missing_in_ad.append(d)
        
        # Liste formatına çevir
        missing_in_system_list = [
            {'gid': gid, 'source_files': files} 
            for gid, files in grouped_missing_in_system.items()
        ]
        
        context['missing_in_system'] = missing_in_system_list
        context['missing_in_ad'] = missing_in_ad
        context['discrepancies'] = all_discrepancies  # Eski template uyumluluğu için
        context['processed_files'] = self.object.processed_files.all()
        return context


class ADLogDeleteView(AdminRequiredMixin, DeleteView):
    """AD Log analizi sil"""
    model = ADLogAnalysis
    template_name = 'it_tools/ad_log_confirm_delete.html'
    success_url = reverse_lazy('it_tools:ad_log_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'AD Log analizi silindi.')
        return super().delete(request, *args, **kwargs)


@login_required
@admin_required
@require_POST
def ad_log_run_analysis(request, pk):
    """AD Log analizini çalıştır"""
    analysis = get_object_or_404(ADLogAnalysis, pk=pk)
    
    # Zaten işlenmiş mi?
    if analysis.status in ['completed', 'email_pending', 'email_sent']:
        return JsonResponse({
            'success': False,
            'error': 'Bu analiz zaten tamamlanmış.'
        })
    
    # Analiz çalıştır
    analysis.status = 'processing'
    analysis.save()
    
    try:
        service = ADLogService(analysis=analysis)
        
        # Kaynak path'i belirle
        if analysis.source_path_config:
            source_path = analysis.source_path_config.source_path
            output_path = analysis.source_path_config.output_path
        else:
            source_path = analysis.source_path
            output_path = None
        
        if not source_path:
            return JsonResponse({
                'success': False,
                'error': 'Kaynak path tanımlanmamış.'
            })
        
        result = service.run_full_analysis(
            source_path=source_path,
            year=analysis.year,
            month=analysis.month,
            output_path=output_path
        )
        
        if result.success:
            analysis.status = 'email_pending'
            analysis.save()
            return JsonResponse({
                'success': True,
                'data': result.data,
                'message': 'Analiz başarıyla tamamlandı.'
            })
        else:
            analysis.status = 'failed'
            analysis.error_message = result.message
            analysis.save()
            return JsonResponse({
                'success': False,
                'error': result.message
            })
    except Exception as e:
        analysis.status = 'failed'
        analysis.error_message = str(e)
        analysis.save()
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
@admin_required
@require_GET
def ad_log_email_preview(request, pk):
    """Email önizlemesi - varsayılan template'i kullanarak"""
    analysis = get_object_or_404(ADLogAnalysis, pk=pk)
    
    # AD_LOG usage type'ını al
    from .models import UsageType
    ad_log_type = UsageType.objects.filter(code='AD_LOG').first()
    
    # Varsayılan AD_LOG template'ini al
    template = None
    if ad_log_type:
        template = ADLogEmailTemplate.objects.filter(
            usage_type=ad_log_type,
            is_default=True,
            is_active=True
        ).first()
    
    if not template and ad_log_type:
        # Template yoksa fallback
        template = ADLogEmailTemplate.objects.filter(
            usage_type=ad_log_type,
            is_active=True
        ).first()
    
    if not template:
        return JsonResponse({
            'success': False,
            'error': 'Email şablonu bulunamadı. Lütfen önce bir email şablonu oluşturun.'
        })
    
    # Eşleşmeyen GID'leri al (missing_in_system)
    unmatched_gids = []
    if analysis.status in ['completed', 'email_pending', 'email_sent']:
        unmatched = GIDDiscrepancy.objects.filter(
            analysis=analysis,
            discrepancy_type='missing_in_system'
        ).values_list('gid', flat=True)[:100]  # İlk 100
        unmatched_gids = list(unmatched)
    
    # Path bilgilerini al
    path_name = ''
    source_path = ''
    output_path = ''
    if analysis.source_path_config:
        path_name = analysis.source_path_config.name
        source_path = analysis.source_path_config.source_path
        output_path = analysis.source_path_config.output_path
    else:
        # Fallback: analysis'ten direkt al
        source_path = analysis.source_path or ''
    
    # Context hazırla
    context = {
        'analysis_name': analysis.name,
        'period': analysis.period_display,
        'date': datetime.now().strftime('%d.%m.%Y'),
        'total_gids': analysis.total_gids_found,
        'unique_gids': analysis.unique_gids_count,
        'discrepancy_count': analysis.discrepancy_count,
        'unmatched_gids': ', '.join(unmatched_gids) if unmatched_gids else 'Yok',
        'path_name': path_name,
        'source_path': source_path,
        'output_path': output_path,
    }
    
    # Template'i render et
    rendered = template.render(context)
    
    return JsonResponse({
        'success': True,
        'data': {
            'subject': rendered['subject'],
            'body': rendered['body'],
            'default_to': rendered['to'],
            'default_cc': rendered['cc']
        }
    })


@login_required
@admin_required
@require_POST
def ad_log_send_email(request, pk):
    """Email gönder"""
    analysis = get_object_or_404(ADLogAnalysis, pk=pk)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Geçersiz veri formatı.'})
    
    to_list = [e.strip() for e in data.get('to', '').split(',') if e.strip()]
    cc_list = [e.strip() for e in data.get('cc', '').split(',') if e.strip()]
    subject = data.get('subject', '')
    body = data.get('body', '')
    
    if not to_list:
        return JsonResponse({'success': False, 'error': 'En az bir alıcı belirtmelisiniz.'})
    
    if not subject:
        return JsonResponse({'success': False, 'error': 'Email konusu boş olamaz.'})
    
    # Ekleri hazırla
    attachments = []
    
    if analysis.log_file:
        attachments.append((
            os.path.basename(analysis.log_file.name),
            analysis.log_file.read(),
            'text/plain'
        ))
    
    if analysis.unique_gids_file:
        attachments.append((
            os.path.basename(analysis.unique_gids_file.name),
            analysis.unique_gids_file.read(),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ))
    
    if analysis.user_checklist_file:
        attachments.append((
            os.path.basename(analysis.user_checklist_file.name),
            analysis.user_checklist_file.read(),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ))
    
    try:
        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=to_list,
            cc=cc_list if cc_list else None
        )
        
        for filename, content, mimetype in attachments:
            email.attach(filename, content, mimetype)
        
        email.send()
        
        analysis.status = 'email_sent'
        analysis.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Email başarıyla gönderildi.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Email gönderilirken hata: {str(e)}'
        })


@login_required
@admin_required
@require_GET
def ad_log_progress(request, pk):
    """
    Server-Sent Events endpoint - Analiz ilerleme durumunu gerçek zamanlı olarak döndürür
    """
    analysis = get_object_or_404(ADLogAnalysis, pk=pk)
    
    def event_stream():
        progress_key = f"ad_log_progress_{pk}"
        last_progress = None
        timeout_counter = 0
        max_timeout = 300  # 5 dakika
        
        while timeout_counter < max_timeout:
            # Cache'den ilerleme verisi al
            current_progress = cache.get(progress_key)
            
            if current_progress and current_progress != last_progress:
                # Yeni ilerleme varsa gönder
                yield f"data: {json.dumps(current_progress)}\n\n"
                last_progress = current_progress
                
                # Eğer tamamlandıysa bağlantıyı kes
                if current_progress.get('step') == 'completed' or current_progress.get('progress') == 100:
                    if current_progress.get('step') in ['saving', 'completed']:
                        break
            
            # Analiz durumunu kontrol et
            analysis.refresh_from_db()
            if analysis.status in ['email_pending', 'completed', 'failed']:
                # İşlem bitti
                final_data = {
                    'step': 'completed',
                    'progress': 100,
                    'message': 'Analiz tamamlandı' if analysis.status != 'failed' else 'Analiz başarısız',
                    'status': analysis.status,
                    'timestamp': datetime.now().isoformat()
                }
                yield f"data: {json.dumps(final_data)}\n\n"
                break
            
            time.sleep(0.5)  # 500ms bekle
            timeout_counter += 0.5
    
    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response


# ===============================
# Şirket Kullanıcıları (Company Users)
# ===============================

class CompanyUserForm(forms.ModelForm):
    """Şirket kullanıcısı form"""
    password1 = forms.CharField(
        label='Şifre',
        widget=forms.PasswordInput(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-indigo-500 focus:border-indigo-500'}),
        required=False,
        help_text='Yeni kullanıcı için zorunlu, düzenlerken boş bırakılabilir.'
    )
    password2 = forms.CharField(
        label='Şifre (Tekrar)',
        widget=forms.PasswordInput(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-indigo-500 focus:border-indigo-500'}),
        required=False
    )
    
    class Meta:
        model = User
        fields = ['username', 'email', 'first_name', 'last_name', 'user_type', 'department', 'gid', 'phone', 'is_active']
        widgets = {
            'username': forms.TextInput(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-indigo-500 focus:border-indigo-500'}),
            'email': forms.EmailInput(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-indigo-500 focus:border-indigo-500'}),
            'first_name': forms.TextInput(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-indigo-500 focus:border-indigo-500'}),
            'last_name': forms.TextInput(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-indigo-500 focus:border-indigo-500'}),
            'user_type': forms.Select(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-indigo-500 focus:border-indigo-500'}),
            'department': forms.Select(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-indigo-500 focus:border-indigo-500'}),
            'gid': forms.TextInput(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-indigo-500 focus:border-indigo-500', 'placeholder': 'Active Directory GID'}),
            'phone': forms.TextInput(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-indigo-500 focus:border-indigo-500'}),
            'is_active': forms.CheckboxInput(attrs={'class': 'h-4 w-4 text-indigo-600 focus:ring-indigo-500 border-gray-300 rounded'}),
        }
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Sadece admin ve salesperson tiplerini göster
        self.fields['user_type'].choices = [
            ('admin', 'Yönetici'),
            ('salesperson', 'Satış Elemanı'),
        ]
    
    def clean(self):
        cleaned_data = super().clean()
        password1 = cleaned_data.get('password1')
        password2 = cleaned_data.get('password2')
        
        # Yeni kullanıcı için şifre zorunlu
        if not self.instance.pk and not password1:
            raise forms.ValidationError('Yeni kullanıcı için şifre zorunludur.')
        
        if password1 and password1 != password2:
            raise forms.ValidationError('Şifreler eşleşmiyor.')
        
        return cleaned_data
    
    def save(self, commit=True):
        user = super().save(commit=False)
        password = self.cleaned_data.get('password1')
        if password:
            user.set_password(password)
        if commit:
            user.save()
        return user


class CompanyUserListView(AdminRequiredMixin, ListView):
    """Şirket kullanıcıları listesi"""
    model = User
    template_name = 'it_tools/company_user_list.html'
    context_object_name = 'users'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = User.objects.exclude(user_type='customer').select_related('department')
        
        # Arama
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(username__icontains=search) |
                models.Q(first_name__icontains=search) |
                models.Q(last_name__icontains=search) |
                models.Q(email__icontains=search) |
                models.Q(gid__icontains=search)
            )
        
        # Departman filtresi
        department = self.request.GET.get('department')
        if department:
            queryset = queryset.filter(department_id=department)
        
        # Kullanıcı tipi filtresi
        user_type = self.request.GET.get('user_type')
        if user_type:
            queryset = queryset.filter(user_type=user_type)
        
        # Aktiflik filtresi
        is_active = self.request.GET.get('is_active')
        if is_active == '1':
            queryset = queryset.filter(is_active=True)
        elif is_active == '0':
            queryset = queryset.filter(is_active=False)
        
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['departments'] = Department.objects.filter(is_active=True)
        return context


class CompanyUserCreateView(AdminRequiredMixin, CreateView):
    """Yeni şirket kullanıcısı oluştur"""
    model = User
    form_class = CompanyUserForm
    template_name = 'it_tools/company_user_form.html'
    success_url = reverse_lazy('it_tools:company_user_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Kullanıcı başarıyla oluşturuldu.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['departments'] = Department.objects.filter(is_active=True)
        return context


class CompanyUserDetailView(AdminRequiredMixin, DetailView):
    """Şirket kullanıcısı detayı"""
    model = User
    template_name = 'it_tools/company_user_detail.html'
    context_object_name = 'user_obj'
    
    def get_queryset(self):
        return User.objects.exclude(user_type='customer')


class CompanyUserUpdateView(AdminRequiredMixin, UpdateView):
    """Şirket kullanıcısı düzenle"""
    model = User
    form_class = CompanyUserForm
    template_name = 'it_tools/company_user_form.html'
    
    def get_queryset(self):
        return User.objects.exclude(user_type='customer')
    
    def get_success_url(self):
        return reverse('it_tools:company_user_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Kullanıcı başarıyla güncellendi.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['departments'] = Department.objects.filter(is_active=True)
        return context


class CompanyUserDeleteView(AdminRequiredMixin, DeleteView):
    """Şirket kullanıcısı sil"""
    model = User
    template_name = 'it_tools/company_user_confirm_delete.html'
    success_url = reverse_lazy('it_tools:company_user_list')
    context_object_name = 'user_obj'
    
    def get_queryset(self):
        return User.objects.exclude(user_type='customer')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Kullanıcı silindi.')
        return super().delete(request, *args, **kwargs)


@login_required
@admin_required
@require_POST
def company_user_toggle_active(request, pk):
    """Kullanıcı aktifliğini değiştir"""
    user = get_object_or_404(User.objects.exclude(user_type='customer'), pk=pk)
    user.is_active = not user.is_active
    user.save()
    status = 'aktif' if user.is_active else 'pasif'
    messages.success(request, f'Kullanıcı {status} yapıldı.')
    return redirect('it_tools:company_user_list')


# ===============================
# Müşteri Kullanıcıları
# ===============================

class CustomerUserListView(AdminRequiredMixin, ListView):
    """Müşteri kullanıcıları listesi"""
    model = Customer
    template_name = 'it_tools/customer_user_list.html'
    context_object_name = 'customers'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = Customer.objects.select_related('user_account', 'salesperson')
        
        # Arama
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(company_name__icontains=search) |
                models.Q(contact_person__icontains=search) |
                models.Q(email__icontains=search) |
                models.Q(tax_number__icontains=search)
            )
        
        # Hesap durumu filtresi
        has_account = self.request.GET.get('has_account')
        if has_account == '1':
            queryset = queryset.filter(user_account__isnull=False)
        elif has_account == '0':
            queryset = queryset.filter(user_account__isnull=True)
        
        return queryset.order_by('-created_at')


class CustomerUserDetailView(AdminRequiredMixin, DetailView):
    """Müşteri detayı"""
    model = Customer
    template_name = 'it_tools/customer_user_detail.html'
    context_object_name = 'customer'


@login_required
@admin_required
@require_POST
def customer_create_account(request, pk):
    """Müşteri için hesap oluştur"""
    customer = get_object_or_404(Customer, pk=pk)
    
    if customer.user_account:
        messages.warning(request, 'Bu müşterinin zaten bir hesabı var.')
        return redirect('it_tools:customer_user_detail', pk=pk)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        # Form data
        data = {
            'username': request.POST.get('username'),
            'password': request.POST.get('password'),
        }
    
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        messages.error(request, 'Kullanıcı adı ve şifre zorunludur.')
        return redirect('it_tools:customer_user_detail', pk=pk)
    
    # Kullanıcı adı kontrolü
    if User.objects.filter(username=username).exists():
        messages.error(request, 'Bu kullanıcı adı zaten kullanılıyor.')
        return redirect('it_tools:customer_user_detail', pk=pk)
    
    # Kullanıcı oluştur
    user = User.objects.create_user(
        username=username,
        password=password,
        email=customer.email,
        first_name=customer.contact_person.split()[0] if customer.contact_person else '',
        last_name=' '.join(customer.contact_person.split()[1:]) if customer.contact_person else '',
        user_type='customer',
        is_active=True
    )
    
    # Müşteriyi kullanıcıya bağla
    customer.user_account = user
    customer.save()
    
    messages.success(request, f'"{username}" hesabı oluşturuldu.')
    return redirect('it_tools:customer_user_detail', pk=pk)


# ===============================
# Email Şablonları
# ===============================

class EmailTemplateListView(AdminRequiredMixin, ListView):
    """Email şablonları listesi"""
    model = ADLogEmailTemplate
    template_name = 'it_tools/email_template_list.html'
    context_object_name = 'email_templates'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = super().get_queryset()
        usage_type = self.request.GET.get('type')
        if usage_type:
            queryset = queryset.filter(usage_type=usage_type)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import UsageType
        context['usage_types'] = UsageType.objects.filter(is_active=True)
        context['selected_type'] = self.request.GET.get('type', '')
        return context


class EmailTemplateCreateView(AdminRequiredMixin, CreateView):
    """Yeni email şablonu oluştur"""
    model = ADLogEmailTemplate
    template_name = 'it_tools/email_template_form.html'
    fields = ['name', 'usage_type', 'subject', 'body', 'default_to', 'default_cc', 'is_active', 'is_default']
    success_url = reverse_lazy('it_tools:email_template_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Email şablonu oluşturuldu.')
        return super().form_valid(form)


class EmailTemplateUpdateView(AdminRequiredMixin, UpdateView):
    """Email şablonu düzenle"""
    model = ADLogEmailTemplate
    template_name = 'it_tools/email_template_form.html'
    fields = ['name', 'usage_type', 'subject', 'body', 'default_to', 'default_cc', 'is_active', 'is_default']
    success_url = reverse_lazy('it_tools:email_template_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Email şablonu güncellendi.')
        return super().form_valid(form)


class EmailTemplateDeleteView(AdminRequiredMixin, DeleteView):
    """Email şablonu sil"""
    model = ADLogEmailTemplate
    template_name = 'it_tools/email_template_confirm_delete.html'
    success_url = reverse_lazy('it_tools:email_template_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Email şablonu silindi.')
        return super().delete(request, *args, **kwargs)


# ===============================
# Toplu Kullanıcı Import
# ===============================

class BulkUserImportForm(forms.ModelForm):
    """Toplu kullanıcı import formu"""
    
    class Meta:
        model = BulkUserImport
        fields = ['name', 'excel_file']
        widgets = {
            'name': forms.TextInput(attrs={
                'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-indigo-500 focus:border-indigo-500',
                'placeholder': 'Örn: Ocak 2026 Kullanıcı Güncellemesi'
            }),
            'excel_file': forms.FileInput(attrs={
                'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-indigo-500 focus:border-indigo-500',
                'accept': '.xlsx,.xls'
            })
        }
    
    def clean_excel_file(self):
        file = self.cleaned_data.get('excel_file')
        if file:
            # Dosya uzantısı kontrolü
            ext = os.path.splitext(file.name)[1].lower()
            if ext not in ['.xlsx', '.xls']:
                raise forms.ValidationError('Sadece Excel dosyaları (.xlsx, .xls) yüklenebilir.')
            
            # Dosya boyutu kontrolü (max 10MB)
            if file.size > 10 * 1024 * 1024:
                raise forms.ValidationError('Dosya boyutu 10MB\'dan büyük olamaz.')
        
        return file


class BulkUserImportListView(AdminRequiredMixin, ListView):
    """Toplu import geçmişi listesi"""
    model = BulkUserImport
    template_name = 'it_tools/bulk_user_import_list.html'
    context_object_name = 'imports'
    paginate_by = 20
    
    def get_queryset(self):
        return BulkUserImport.objects.all().order_by('-created_at')


class BulkUserImportCreateView(AdminRequiredMixin, CreateView):
    """Yeni toplu import oluştur"""
    model = BulkUserImport
    form_class = BulkUserImportForm
    template_name = 'it_tools/bulk_user_import_form.html'
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        
        # Import işlemini başlat
        process_bulk_import(self.object.pk)
        
        messages.success(self.request, 'Import işlemi başlatıldı.')
        return response
    
    def get_success_url(self):
        return reverse('it_tools:bulk_user_import_detail', kwargs={'pk': self.object.pk})


class BulkUserImportDetailView(AdminRequiredMixin, DetailView):
    """Import detayı"""
    model = BulkUserImport
    template_name = 'it_tools/bulk_user_import_detail.html'
    context_object_name = 'import_obj'


@login_required
@admin_required
def bulk_user_import_download_sample(request):
    """Örnek Excel dosyası indir"""
    import openpyxl
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Users'
    
    # Header'ları ekle
    headers = [
        'Title', 'Academic degrees', 'Surname', 'Given name',
        'Surname (national)', 'Given name (national)', 'Nickname', 'GID',
        'Function', 'Department (org code)', 'Department (long text)',
        'Country', 'Location', 'Organisation', 'Company', 'Company Name',
        'Building', 'Room number', 'Telephone number', 'Alternate phone number',
        'Mobile phone number', 'E-Mail', 'Cost center', 'ARE/AUN',
        'CostLocUnitName', 'OrgID', 'Secretary', 'Representation',
        'Sponsor', 'Manager', 'Record type', 'User type', 'Status',
        'Letterbox', 'Contract status', 'Properties'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        # Sütun genişliği ayarla
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Örnek satır ekle
    sample_data = [
        'Mr.', 'Dr.', 'Yılmaz', 'Ahmet',
        'Yılmaz', 'Ahmet', 'ahmety', 'GID123456',
        'Manager', 'SALES01', 'Sales Department',
        'Turkey', 'Istanbul', 'Corp', 'ABC', 'ABC Company',
        'HQ', '101', '+90 212 555 1234', '+90 212 555 5678',
        '+90 532 555 9012', 'ahmet.yilmaz@company.com', 'CC001', 'AUN01',
        'Sales Unit', 'ORG001', '', '',
        '', 'manager.name@company.com', 'Employee', 'Internal', 'Active',
        '', 'Permanent', ''
    ]
    
    for col, value in enumerate(sample_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    # Response oluştur
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=kullanici_import_sablonu.xlsx'
    
    return response


# ===============================
# Departman Yönetimi
# ===============================

class DepartmentForm(forms.ModelForm):
    """Departman formu"""
    
    class Meta:
        model = Department
        fields = ['name', 'code', 'org_code', 'description', 'is_active']
        widgets = {
            'name': forms.TextInput(attrs={
                'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-indigo-500 focus:border-indigo-500',
                'placeholder': 'Departman adı'
            }),
            'code': forms.TextInput(attrs={
                'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-indigo-500 focus:border-indigo-500',
                'placeholder': 'Otomatik oluşturulur (opsiyonel)'
            }),
            'org_code': forms.TextInput(attrs={
                'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-indigo-500 focus:border-indigo-500',
                'placeholder': 'Excel org code (opsiyonel)'
            }),
            'description': forms.Textarea(attrs={
                'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-indigo-500 focus:border-indigo-500',
                'rows': 3,
                'placeholder': 'Açıklama (opsiyonel)'
            }),
            'is_active': forms.CheckboxInput(attrs={
                'class': 'h-4 w-4 text-indigo-600 focus:ring-indigo-500 border-gray-300 rounded'
            })
        }


class DepartmentListView(AdminRequiredMixin, ListView):
    """Departman listesi"""
    model = Department
    template_name = 'it_tools/department_list.html'
    context_object_name = 'departments'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = Department.objects.all()
        
        # Arama
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(name__icontains=search) |
                models.Q(code__icontains=search) |
                models.Q(org_code__icontains=search)
            )
        
        # Aktiflik filtresi
        is_active = self.request.GET.get('is_active')
        if is_active == '1':
            queryset = queryset.filter(is_active=True)
        elif is_active == '0':
            queryset = queryset.filter(is_active=False)
        
        return queryset.order_by('name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_departments'] = Department.objects.count()
        context['active_departments'] = Department.objects.filter(is_active=True).count()
        return context


class DepartmentCreateView(AdminRequiredMixin, CreateView):
    """Yeni departman oluştur"""
    model = Department
    form_class = DepartmentForm
    template_name = 'it_tools/department_form.html'
    success_url = reverse_lazy('it_tools:department_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Departman başarıyla oluşturuldu.')
        return super().form_valid(form)


class DepartmentUpdateView(AdminRequiredMixin, UpdateView):
    """Departman düzenle"""
    model = Department
    form_class = DepartmentForm
    template_name = 'it_tools/department_form.html'
    success_url = reverse_lazy('it_tools:department_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Departman başarıyla güncellendi.')
        return super().form_valid(form)


class DepartmentDetailView(AdminRequiredMixin, DetailView):
    """Departman detayı"""
    model = Department
    template_name = 'it_tools/department_detail.html'
    context_object_name = 'department'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['users'] = self.object.users.all()[:20]
        context['total_users'] = self.object.users.count()
        return context


class DepartmentDeleteView(AdminRequiredMixin, DeleteView):
    """Departman sil"""
    model = Department
    template_name = 'it_tools/department_confirm_delete.html'
    success_url = reverse_lazy('it_tools:department_list')
    
    def delete(self, request, *args, **kwargs):
        dept = self.get_object()
        if dept.users.exists():
            messages.error(request, f'Bu departmanda {dept.users.count()} kullanıcı var. Önce kullanıcıları başka departmana taşıyın.')
            return redirect('it_tools:department_detail', pk=dept.pk)
        messages.success(request, 'Departman silindi.')
        return super().delete(request, *args, **kwargs)


@login_required
@admin_required
@require_POST
def department_toggle_active(request, pk):
    """Departman aktifliğini değiştir"""
    dept = get_object_or_404(Department, pk=pk)
    dept.is_active = not dept.is_active
    dept.save()
    status = 'aktif' if dept.is_active else 'pasif'
    messages.success(request, f'Departman {status} yapıldı.')
    return redirect('it_tools:department_list')
