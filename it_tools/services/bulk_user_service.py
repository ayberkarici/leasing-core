"""
Toplu kullanıcı import servisi
Excel dosyasından kullanıcıları sisteme ekler veya günceller
"""

import openpyxl
from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils import timezone
from datetime import datetime
import logging

from accounts.models import Department

User = get_user_model()
logger = logging.getLogger(__name__)


# Excel header'ları ve model field eşleştirmesi
EXCEL_HEADER_MAPPING = {
    'Title': 'title',
    'Academic degrees': 'academic_degrees',
    'Surname': 'last_name',
    'Given name': 'first_name',
    'Surname (national)': 'surname_national',
    'Given name (national)': 'given_name_national',
    'Nickname': 'nickname',
    'GID': 'gid',
    'Function': 'function',
    'Department (org code)': 'department_org_code',
    'Department (long text)': 'department_long_text',
    'Country': 'country',
    'Location': 'location',
    'Organisation': 'organisation',
    'Company': 'company',
    'Company Name': 'company_name',
    'Building': 'building',
    'Room number': 'room_number',
    'Telephone number': 'phone',
    'Alternate phone number': 'alternate_phone',
    'Mobile phone number': 'mobile_phone',
    'E-Mail': 'email',
    'Cost center': 'cost_center',
    'ARE/AUN': 'are_aun',
    'CostLocUnitName': 'cost_loc_unit_name',
    'OrgID': 'org_id',
    'Secretary': 'secretary',
    'Representation': 'representation',
    'Sponsor': 'sponsor',
    'Manager': 'manager',
    'Record type': 'record_type',
    'User type': 'excel_user_type',
    'Status': 'excel_status',
    'Letterbox': 'letterbox',
    'Contract status': 'contract_status',
    'Properties': 'properties',
}


class BulkUserImportService:
    """Toplu kullanıcı import servisi"""
    
    def __init__(self, import_record):
        self.import_record = import_record
        self.logs = []
        self.errors = []
        self.created_count = 0
        self.updated_count = 0
        self.skipped_count = 0
        self.error_count = 0
    
    def log(self, message):
        """Log mesajı ekle"""
        timestamp = timezone.now().strftime('%H:%M:%S')
        self.logs.append(f"[{timestamp}] {message}")
    
    def error(self, row_num, message):
        """Hata mesajı ekle"""
        self.errors.append(f"Satır {row_num}: {message}")
        self.error_count += 1
    
    def process(self):
        """Excel dosyasını işle"""
        try:
            self.import_record.status = 'processing'
            self.import_record.save()
            
            self.log("İşlem başlatıldı...")
            
            # Excel dosyasını aç
            wb = openpyxl.load_workbook(self.import_record.excel_file.path)
            ws = wb.active
            
            # Header'ları al
            headers = [cell.value for cell in ws[1]]
            self.log(f"Bulunan sütunlar: {len(headers)}")
            
            # Header mapping oluştur
            header_index = {}
            for idx, header in enumerate(headers):
                if header and header.strip() in EXCEL_HEADER_MAPPING:
                    header_index[EXCEL_HEADER_MAPPING[header.strip()]] = idx
            
            self.log(f"Eşleştirilen sütunlar: {len(header_index)}")
            
            # GID sütunu zorunlu
            if 'gid' not in header_index:
                self.error(1, "GID sütunu bulunamadı! Import iptal edildi.")
                self.import_record.status = 'failed'
                self.import_record.error_details = '\n'.join(self.errors)
                self.import_record.log = '\n'.join(self.logs)
                self.import_record.save()
                return False
            
            # Satırları işle
            total_rows = ws.max_row - 1  # Header hariç
            self.import_record.total_rows = total_rows
            self.log(f"Toplam {total_rows} satır işlenecek")
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    self._process_row(row_num, row, header_index)
                except Exception as e:
                    self.error(row_num, f"Beklenmeyen hata: {str(e)}")
                    logger.exception(f"Row {row_num} processing error")
            
            # Sonuçları kaydet
            self.import_record.created_count = self.created_count
            self.import_record.updated_count = self.updated_count
            self.import_record.skipped_count = self.skipped_count
            self.import_record.error_count = self.error_count
            self.import_record.log = '\n'.join(self.logs)
            self.import_record.error_details = '\n'.join(self.errors) if self.errors else None
            self.import_record.completed_at = timezone.now()
            
            if self.error_count > 0:
                self.import_record.status = 'completed_with_errors'
            else:
                self.import_record.status = 'completed'
            
            self.import_record.save()
            
            self.log(f"İşlem tamamlandı: {self.created_count} oluşturuldu, {self.updated_count} güncellendi, {self.skipped_count} atlandı, {self.error_count} hata")
            
            return True
            
        except Exception as e:
            logger.exception("Bulk import failed")
            self.import_record.status = 'failed'
            self.import_record.error_details = f"Kritik hata: {str(e)}"
            self.import_record.log = '\n'.join(self.logs)
            self.import_record.save()
            return False
    
    def _process_row(self, row_num, row, header_index):
        """Tek bir satırı işle"""
        
        def get_value(field):
            """Belirtilen field için değeri al"""
            if field in header_index:
                val = row[header_index[field]]
                if val is not None:
                    return str(val).strip()
            return None
        
        # GID zorunlu
        gid = get_value('gid')
        if not gid:
            self.skipped_count += 1
            return
        
        # Email zorunlu (username olarak kullanılacak)
        email = get_value('email')
        if not email:
            self.error(row_num, f"GID {gid}: Email adresi bulunamadı")
            return
        
        # Ad ve soyad
        first_name = get_value('first_name') or ''
        last_name = get_value('last_name') or ''
        
        if not first_name and not last_name:
            self.error(row_num, f"GID {gid}: Ad veya soyad bulunamadı")
            return
        
        # Username - email'den @ öncesini al
        username = email.split('@')[0].lower()
        
        # Kullanıcıyı bul veya oluştur (GID'ye göre)
        try:
            with transaction.atomic():
                user, created = User.objects.get_or_create(
                    gid=gid,
                    defaults={
                        'username': username,
                        'email': email,
                        'first_name': first_name,
                        'last_name': last_name,
                        'user_type': 'salesperson',  # Varsayılan olarak satış elemanı
                        'is_active': True,
                    }
                )
                
                if created:
                    self.created_count += 1
                    self.log(f"Yeni kullanıcı oluşturuldu: {gid} - {first_name} {last_name}")
                else:
                    # Mevcut kullanıcıyı güncelle
                    updated = False
                    
                    if user.email != email:
                        user.email = email
                        updated = True
                    
                    if user.first_name != first_name:
                        user.first_name = first_name
                        updated = True
                    
                    if user.last_name != last_name:
                        user.last_name = last_name
                        updated = True
                    
                    # Telefon numarası
                    phone = get_value('phone')
                    if phone and user.phone != phone:
                        user.phone = phone
                        updated = True
                    
                    if updated:
                        user.save()
                        self.updated_count += 1
                        self.log(f"Kullanıcı güncellendi: {gid} - {first_name} {last_name}")
                    else:
                        self.skipped_count += 1
                
                # Departman eşleştirmesi - her zaman güncelle
                dept_name = get_value('department_long_text')
                dept_org_code = get_value('department_org_code')
                
                if dept_name:
                    dept = self._get_or_create_department(dept_name, dept_org_code)
                    if dept and user.department != dept:
                        user.department = dept
                        user.save()
                        if not created:
                            self.log(f"Kullanıcı departmanı güncellendi: {gid} -> {dept.name}")
                
        except Exception as e:
            self.error(row_num, f"GID {gid}: Kayıt hatası - {str(e)}")
    
    def _get_or_create_department(self, dept_name, org_code=None):
        """Departman adına göre bul veya oluştur"""
        if not dept_name:
            return None
        
        dept_name = dept_name.strip()
        
        # Önce isimle ara
        try:
            dept = Department.objects.get(name__iexact=dept_name)
            # Org code varsa güncelle
            if org_code and not dept.org_code:
                dept.org_code = org_code
                dept.save()
            return dept
        except Department.DoesNotExist:
            pass
        
        # Org code ile ara (varsa)
        if org_code:
            try:
                dept = Department.objects.get(org_code=org_code)
                return dept
            except Department.DoesNotExist:
                pass
        
        # Yeni departman oluştur
        try:
            dept = Department.objects.create(
                name=dept_name,
                org_code=org_code,
                is_active=True
            )
            self.log(f"Yeni departman oluşturuldu: {dept_name}")
            return dept
        except Exception as e:
            self.error(0, f"Departman oluşturulamadı ({dept_name}): {str(e)}")
            return None


def process_bulk_import(import_id):
    """Import işlemini başlat"""
    from it_tools.models import BulkUserImport
    
    try:
        import_record = BulkUserImport.objects.get(pk=import_id)
        service = BulkUserImportService(import_record)
        return service.process()
    except BulkUserImport.DoesNotExist:
        logger.error(f"Import record not found: {import_id}")
        return False
