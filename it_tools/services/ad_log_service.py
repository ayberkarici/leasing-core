import os
import glob
import shutil
import calendar
from datetime import datetime
from typing import List, Dict, Set, Optional, Tuple
from dataclasses import dataclass

from django.conf import settings
from django.core.files.base import ContentFile
from django.core.mail import EmailMessage
from django.template import Template, Context
from django.contrib.auth import get_user_model
from django.core.cache import cache

from core.services.base import BaseService, ServiceResult

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


User = get_user_model()


@dataclass
class GIDRecord:
    """GID kaydı"""
    gid: str
    display_name: Optional[str] = None
    email: Optional[str] = None
    department: Optional[str] = None
    source_file: Optional[str] = None
    date: Optional[str] = None


class ADLogService(BaseService):
    """AD Log analiz servisi"""
    
    # Excel dosya adı formatı: EventExport_YYYY-MM-DD
    FILE_PREFIX = "EventExport_"
    
    def __init__(self, analysis=None):
        super().__init__()
        self.analysis = analysis
        self._gids_from_files: Set[str] = set()
        self._gid_records: List[GIDRecord] = []
        self._processed_files: List[str] = []
        self._temp_dir: Optional[str] = None
        self._progress_key = f"ad_log_progress_{analysis.pk}" if analysis else None
    
    def get_temp_directory(self) -> str:
        """Geçici klasör yolunu döndür"""
        if self._temp_dir:
            return self._temp_dir
        
        temp_base = os.path.join(settings.MEDIA_ROOT, 'ad_logs', 'temp')
        os.makedirs(temp_base, exist_ok=True)
        
        # Analiz için özel temp klasörü
        if self.analysis:
            self._temp_dir = os.path.join(temp_base, f"analysis_{self.analysis.pk}")
        else:
            self._temp_dir = os.path.join(temp_base, f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        
        os.makedirs(self._temp_dir, exist_ok=True)
        return self._temp_dir
    
    def cleanup_temp_directory(self):
        """Geçici klasörü temizle"""
        if self._temp_dir and os.path.exists(self._temp_dir):
            try:
                shutil.rmtree(self._temp_dir)
            except Exception as e:
                self.logger.warning(f"Temp klasör temizlenirken hata: {e}")
    
    def update_progress(self, step: str, progress: int, message: str, details: dict = None):
        """İlerleme durumunu güncelle"""
        if not self._progress_key:
            return
        
        progress_data = {
            'step': step,
            'progress': progress,
            'message': message,
            'details': details or {},
            'timestamp': datetime.now().isoformat()
        }
        cache.set(self._progress_key, progress_data, timeout=3600)  # 1 saat
    
    def get_expected_filenames(self, year: int, month: int) -> List[str]:
        """
        Belirli bir yıl ve ay için beklenen dosya adlarını döndür
        Format: EventExport_YYYY-MM-DD
        """
        filenames = []
        _, days_in_month = calendar.monthrange(year, month)
        
        for day in range(1, days_in_month + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            filename = f"{self.FILE_PREFIX}{date_str}"
            filenames.append(filename)
        
        return filenames
    
    def download_files_to_temp(self, source_path: str, year: int, month: int) -> ServiceResult:
        """
        Kaynak path'ten seçili yıl/ay'a ait dosyaları temp klasöre indir
        """
        self.update_progress('downloading', 0, 'Dosyalar fileserver\'dan indiriliyor...', {'phase': 'start'})
        
        if not os.path.exists(source_path):
            return self.error(f"Kaynak klasör bulunamadı: {source_path}")
        
        temp_dir = self.get_temp_directory()
        expected_files = self.get_expected_filenames(year, month)
        
        downloaded_files = []
        not_found_files = []
        total_files = len(expected_files)
        
        for idx, filename_base in enumerate(expected_files, 1):
            # .xlsx, .xls uzantılarını dene
            found = False
            for ext in ['.xlsx', '.xls', '.xlsm']:
                source_file = os.path.join(source_path, filename_base + ext)
                if os.path.exists(source_file):
                    # Dosyayı temp klasöre kopyala
                    dest_file = os.path.join(temp_dir, filename_base + ext)
                    try:
                        shutil.copy2(source_file, dest_file)
                        downloaded_files.append(dest_file)
                        found = True
                        # Progress güncelle
                        progress = int((idx / total_files) * 100)
                        self.update_progress('downloading', progress, 
                                           f'{idx}/{total_files} dosya indirildi',
                                           {'downloaded': len(downloaded_files), 'missing': len(not_found_files)})
                        break
                    except Exception as e:
                        self.logger.error(f"Dosya kopyalanırken hata: {source_file} - {e}")
            
            if not found:
                not_found_files.append(filename_base)
        
        if not downloaded_files:
            return self.error(f"Seçilen dönem için ({month:02d}/{year}) hiç dosya bulunamadı.")
        
        self.update_progress('downloading', 100, 'Dosyalar başarıyla indirildi',
                           {'downloaded': len(downloaded_files), 'missing': len(not_found_files)})
        
        return self.success({
            'downloaded_count': len(downloaded_files),
            'not_found_count': len(not_found_files),
            'downloaded_files': downloaded_files,
            'not_found_files': not_found_files[:10],  # İlk 10'u göster
            'temp_directory': temp_dir
        })
    
    def process_downloaded_files(self) -> ServiceResult:
        """
        Temp klasördeki Excel dosyalarını işle
        D sütunundaki (MatchedQueryElements) GID'leri çıkar
        """
        self.update_progress('processing', 0, 'Excel dosyaları işleniyor...')
        
        if not OPENPYXL_AVAILABLE and not PANDAS_AVAILABLE:
            return self.error("Excel işleme için openpyxl veya pandas kütüphanesi gerekli.")
        
        temp_dir = self.get_temp_directory()
        
        if not os.path.exists(temp_dir):
            return self.error("Temp klasör bulunamadı. Önce dosyaları indirin.")
        
        # Excel dosyalarını bul
        excel_files = []
        for ext in ['*.xlsx', '*.xls', '*.xlsm']:
            excel_files.extend(glob.glob(os.path.join(temp_dir, ext)))
        
        if not excel_files:
            return self.error(f"Temp klasörde Excel dosyası bulunamadı: {temp_dir}")
        
        processed_count = 0
        total_gids = 0
        
        for file_path in excel_files:
            try:
                gids = self._extract_gids_from_column_d(file_path)
                if gids:
                    self._gids_from_files.update([g.gid for g in gids])
                    self._gid_records.extend(gids)
                    self._processed_files.append(file_path)
                    total_gids += len(gids)
                    processed_count += 1
            except Exception as e:
                self.logger.warning(f"Dosya işlenirken hata: {file_path} - {str(e)}")
                continue
        
        self.update_progress('processing', 100, 'Dosyalar başarıyla işlendi',
                           {'processed': processed_count, 'unique_gids': len(self._gids_from_files)})
        
        return self.success({
            'processed_files': processed_count,
            'total_gids': total_gids,
            'unique_gids': len(self._gids_from_files),
            'files': [os.path.basename(f) for f in self._processed_files]
        })
    
    def _extract_gids_from_column_d(self, file_path: str) -> List[GIDRecord]:
        """
        Excel dosyasının D sütunundaki (MatchedQueryElements) GID'leri çıkar
        """
        gids = []
        filename = os.path.basename(file_path)
        
        # Dosya adından tarih çıkar
        date_str = None
        if self.FILE_PREFIX in filename:
            date_part = filename.replace(self.FILE_PREFIX, '').split('.')[0]
            date_str = date_part
        
        if PANDAS_AVAILABLE:
            try:
                df = pd.read_excel(file_path)
                
                # D sütunu = index 3 (0-indexed) veya sütun adı "MatchedQueryElements"
                gid_column = None
                
                # Sütun adıyla bul
                for col in df.columns:
                    if col and 'matchedqueryelements' in str(col).lower():
                        gid_column = col
                        break
                
                # Bulunamadıysa D sütununu (index 3) dene
                if gid_column is None and len(df.columns) >= 4:
                    gid_column = df.columns[3]
                
                if gid_column:
                    for _, row in df.iterrows():
                        gid_value = row.get(gid_column)
                        if gid_value and pd.notna(gid_value):
                            gid_str = str(gid_value).strip()
                            if gid_str:
                                record = GIDRecord(
                                    gid=gid_str,
                                    source_file=filename,
                                    date=date_str
                                )
                                gids.append(record)
                                
            except Exception as e:
                self.logger.error(f"Pandas ile dosya okunurken hata: {e}")
        
        elif OPENPYXL_AVAILABLE:
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                
                # D sütunu = index 4 (1-indexed)
                for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
                    if row[0]:
                        gid_str = str(row[0]).strip()
                        if gid_str:
                            record = GIDRecord(
                                gid=gid_str,
                                source_file=filename,
                                date=date_str
                            )
                            gids.append(record)
                            
            except Exception as e:
                self.logger.error(f"Openpyxl ile dosya okunurken hata: {e}")
        
        return gids
    
    def compare_with_user_gids(self) -> ServiceResult:
        """
        Excel'den çıkarılan GID'leri kullanıcı GID'leri ile karşılaştır
        Sadece eşleşmeyenleri (sistemde olmayanları) döndür
        """
        self.update_progress('comparing', 0, 'Kullanıcı GID\'leri karşılaştırılıyor...')
        
        if not self._gids_from_files:
            return self.error("Önce Excel dosyalarını işlemelisiniz.")
        
        # Kullanıcıların GID'lerini al (CustomUser modelindeki gid alanı)
        user_gids = set(
            User.objects.exclude(gid__isnull=True)
            .exclude(gid='')
            .values_list('gid', flat=True)
        )
        
        self.update_progress('comparing', 30, 'Sistem GID\'leri alındı...')
        
        # Excel'de olup kullanıcılarda olmayan GID'ler
        unmatched_gids = self._gids_from_files - user_gids
        
        self.update_progress('comparing', 60, 'Farklılıklar tespit ediliyor...')
        
        # Farklılık detaylarını oluştur
        discrepancies = []
        for gid in unmatched_gids:
            # Kaynak dosyayı bul
            sources = set(r.source_file for r in self._gid_records if r.gid == gid)
            discrepancies.append({
                'gid': gid,
                'type': 'not_in_users',
                'details': f"GID '{gid}' AD loglarında mevcut ancak sistemde tanımlı kullanıcı yok.",
                'source_files': list(sources)
            })
        
        # Analiz varsa farklılıkları kaydet
        if self.analysis:
            from it_tools.models import GIDDiscrepancy
            GIDDiscrepancy.objects.filter(analysis=self.analysis).delete()
            for d in discrepancies:
                GIDDiscrepancy.objects.create(
                    analysis=self.analysis,
                    gid=d['gid'],
                    discrepancy_type='missing_in_system',
                    details=d['details'],
                    source_file=', '.join(d['source_files']) if d['source_files'] else None
                )
        
        self.update_progress('comparing', 100, 'Karşılaştırma tamamlandı',
                           {'unmatched': len(unmatched_gids), 'matched': len(self._gids_from_files) - len(unmatched_gids)})
        
        return self.success({
            'total_in_files': len(self._gids_from_files),
            'total_user_gids': len(user_gids),
            'unmatched_count': len(unmatched_gids),
            'matched_count': len(self._gids_from_files) - len(unmatched_gids),
            'unmatched_gids': list(unmatched_gids),
            'discrepancies': discrepancies
        })
    
    def save_outputs(self, output_base_path: str = None) -> ServiceResult:
        """
        Tüm çıktıları belirtilen klasöre kaydet:
        - İndirilen Excel'ler
        - Karşılaştırılan GID'ler (bizimkiler)
        - Eşleşmeyenler log dosyası
        """
        self.update_progress('saving', 0, 'Çıktılar kaydediliyor...')
        
        if not self.analysis:
            return self.error("Analiz objesi gerekli.")
        
        # Çıktı klasörü
        if output_base_path:
            output_folder = output_base_path
        else:
            output_folder = self.analysis.get_output_folder()
        
        os.makedirs(output_folder, exist_ok=True)
        
        saved_files = {}
        
        try:
            # 1. İndirilen Excel'leri kaydet (temp'ten output'a taşı)
            self.update_progress('saving', 20, 'Excel dosyaları kaydediliyor...')
            temp_dir = self.get_temp_directory()
            excel_folder = os.path.join(output_folder, 'excels')
            os.makedirs(excel_folder, exist_ok=True)
            
            for file_path in glob.glob(os.path.join(temp_dir, '*.xls*')):
                dest = os.path.join(excel_folder, os.path.basename(file_path))
                shutil.copy2(file_path, dest)
            saved_files['excel_folder'] = excel_folder
            
            # 2. Kullanıcı GID'lerini Excel olarak kaydet
            self.update_progress('saving', 40, 'Kullanıcı GID\'leri kaydediliyor...')
            user_gids_result = self._save_user_gids_excel(output_folder)
            if user_gids_result.success:
                saved_files['user_gids_file'] = user_gids_result.data['filepath']
            
            # 3. Unique GID'leri kaydet (Excel'lerden çekilen)
            self.update_progress('saving', 60, 'Unique GID\'ler kaydediliyor...')
            unique_gids_result = self._save_unique_gids_excel(output_folder)
            if unique_gids_result.success:
                saved_files['unique_gids_file'] = unique_gids_result.data['filepath']
            
            # 4. Eşleşmeyenleri log dosyası olarak kaydet
            self.update_progress('saving', 80, 'Farklılık raporu kaydediliyor...')
            log_result = self._save_discrepancy_log(output_folder)
            if log_result.success:
                saved_files['log_file'] = log_result.data['filepath']
            
            self.update_progress('saving', 100, 'Tüm çıktılar kaydedildi')
            
            return self.success({
                'output_folder': output_folder,
                'saved_files': saved_files
            })
            
        except Exception as e:
            return self.error(f"Dosyalar kaydedilirken hata: {str(e)}")
    
    def _save_user_gids_excel(self, output_folder: str) -> ServiceResult:
        """Kullanıcı GID'lerini Excel olarak kaydet"""
        if not OPENPYXL_AVAILABLE:
            return self.error("openpyxl gerekli")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Kullanıcı GID'leri"
        
        # Başlıklar
        headers = ['GID', 'Ad Soyad', 'Email', 'Departman', 'Kullanıcı Tipi', 'Aktif']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Kullanıcı verileri
        users = User.objects.exclude(gid__isnull=True).exclude(gid='').select_related('department')
        
        for row_num, user in enumerate(users, 2):
            ws.cell(row=row_num, column=1, value=user.gid)
            ws.cell(row=row_num, column=2, value=user.full_name)
            ws.cell(row=row_num, column=3, value=user.email)
            ws.cell(row=row_num, column=4, value=user.department.name if user.department else '')
            ws.cell(row=row_num, column=5, value=user.get_user_type_display())
            ws.cell(row=row_num, column=6, value='Evet' if user.is_active else 'Hayır')
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 10
        
        filepath = os.path.join(output_folder, 'kullanici_gidleri.xlsx')
        wb.save(filepath)
        
        return self.success({'filepath': filepath})
    
    def _save_unique_gids_excel(self, output_folder: str) -> ServiceResult:
        """Excel'lerden çekilen unique GID'leri kaydet"""
        if not OPENPYXL_AVAILABLE:
            return self.error("openpyxl gerekli")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "AD Log GID'leri"
        
        # Başlıklar
        headers = ['GID', 'Kaynak Dosyalar']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # GID verileri
        gid_sources = {}
        for record in self._gid_records:
            if record.gid not in gid_sources:
                gid_sources[record.gid] = set()
            if record.source_file:
                gid_sources[record.gid].add(record.source_file)
        
        for row_num, (gid, sources) in enumerate(sorted(gid_sources.items()), 2):
            ws.cell(row=row_num, column=1, value=gid)
            ws.cell(row=row_num, column=2, value=', '.join(sources))
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 80
        
        filepath = os.path.join(output_folder, 'ad_log_gidleri.xlsx')
        wb.save(filepath)
        
        return self.success({'filepath': filepath})
    
    def _save_discrepancy_log(self, output_folder: str) -> ServiceResult:
        """Eşleşmeyen GID'leri log dosyası olarak kaydet"""
        from it_tools.models import GIDDiscrepancy
        
        if not self.analysis:
            return self.error("Analiz gerekli")
        
        discrepancies = GIDDiscrepancy.objects.filter(analysis=self.analysis)
        
        now = datetime.now()
        filename = f"log_{self.analysis.year}_{self.analysis.month:02d}_{now.strftime('%d%m%y_%H%M')}.txt"
        
        lines = [
            "=" * 70,
            "AD LOG ANALİZ RAPORU - EŞLEŞMEYEN GID'LER",
            "=" * 70,
            "",
            f"Analiz Adı      : {self.analysis.name}",
            f"Dönem           : {self.analysis.period_display}",
            f"Oluşturulma     : {now.strftime('%d/%m/%Y %H:%M:%S')}",
            f"Oluşturan       : {self.analysis.created_by.full_name if self.analysis.created_by else 'N/A'}",
            "",
            "-" * 70,
            "ÖZET",
            "-" * 70,
            f"AD Log'lardan Çekilen Unique GID   : {self.analysis.unique_gids_count}",
            f"Sistemdeki Kullanıcı GID Sayısı    : {User.objects.exclude(gid__isnull=True).exclude(gid='').count()}",
            f"Eşleşmeyen GID Sayısı              : {discrepancies.count()}",
            "",
            "-" * 70,
            "EŞLEŞMEYEN GID'LER (AD'de var, Sistemde yok)",
            "-" * 70,
        ]
        
        if discrepancies.exists():
            for i, d in enumerate(discrepancies, 1):
                lines.append(f"{i:4d}. {d.gid}")
                if d.source_file:
                    lines.append(f"       Kaynak: {d.source_file}")
        else:
            lines.append("  Tüm GID'ler sistemdeki kullanıcılarla eşleşti!")
        
        lines.extend([
            "",
            "=" * 70,
            "RAPOR SONU",
            "=" * 70,
        ])
        
        filepath = os.path.join(output_folder, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        return self.success({'filepath': filepath})
    
    def run_full_analysis(self, source_path: str, year: int, month: int, output_path: str = None) -> ServiceResult:
        """
        Tam analiz sürecini çalıştır:
        1. Dosyaları temp'e indir
        2. GID'leri çıkar
        3. Kullanıcılarla karşılaştır
        4. Çıktıları kaydet
        """
        if not self.analysis:
            return self.error("Analiz objesi gerekli.")
        
        try:
            # Başlangıç progress'i
            self.update_progress('initializing', 0, 'Analiz başlatılıyor...')
            
            # 1. Durumu güncelle
            self.analysis.status = 'downloading'
            self.analysis.save()
            
            # 2. Dosyaları indir
            download_result = self.download_files_to_temp(source_path, year, month)
            if not download_result.success:
                self.analysis.status = 'failed'
                self.analysis.error_message = download_result.message
                self.analysis.save()
                return download_result
            
            # 3. Durumu güncelle
            self.analysis.status = 'processing'
            self.analysis.save()
            
            # 4. Dosyaları işle
            process_result = self.process_downloaded_files()
            if not process_result.success:
                self.analysis.status = 'failed'
                self.analysis.error_message = process_result.message
                self.analysis.save()
                return process_result
            
            # 5. Durumu güncelle
            self.analysis.status = 'comparing'
            self.analysis.processed_files_count = process_result.data['processed_files']
            self.analysis.total_gids_found = process_result.data['total_gids']
            self.analysis.unique_gids_count = process_result.data['unique_gids']
            self.analysis.save()
            
            # 6. Karşılaştır
            compare_result = self.compare_with_user_gids()
            if not compare_result.success:
                self.analysis.status = 'failed'
                self.analysis.error_message = compare_result.message
                self.analysis.save()
                return compare_result
            
            self.analysis.discrepancy_count = compare_result.data['unmatched_count']
            self.analysis.save()
            
            # 7. Çıktıları kaydet
            save_result = self.save_outputs(output_path)
            if not save_result.success:
                self.analysis.status = 'failed'
                self.analysis.error_message = save_result.message
                self.analysis.save()
                return save_result
            
            # 8. Tamamlandı
            self.analysis.status = 'completed'
            self.analysis.error_message = None
            self.analysis.save()
            
            # Final progress
            self.update_progress('completed', 100, 'Analiz başarıyla tamamlandı!', 
                               {'status': 'completed'})
            
            # 9. Temp klasörü temizle
            self.cleanup_temp_directory()
            
            return self.success({
                'download': download_result.data,
                'process': process_result.data,
                'compare': compare_result.data,
                'output': save_result.data
            })
            
        except Exception as e:
            self.analysis.status = 'failed'
            self.analysis.error_message = str(e)
            self.analysis.save()
            return self.error(f"Analiz sırasında hata: {str(e)}")
    
    # ==========================================
    # Email İşlemleri
    # ==========================================
    
    def send_email(self, to_list: List[str], cc_list: List[str], subject: str, 
                   body: str, attachments: List[Tuple] = None) -> ServiceResult:
        """
        Analiz sonuçlarını email olarak gönder
        """
        try:
            email = EmailMessage(
                subject=subject,
                body=body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=to_list,
                cc=cc_list
            )
            
            if attachments:
                for filename, content, mimetype in attachments:
                    email.attach(filename, content, mimetype)
            
            email.send()
            
            # Analiz email bilgilerini güncelle
            if self.analysis:
                self.analysis.email_to = ', '.join(to_list)
                self.analysis.email_cc = ', '.join(cc_list) if cc_list else None
                self.analysis.email_subject = subject
                self.analysis.email_body = body
                self.analysis.email_sent_at = datetime.now()
                self.analysis.status = 'email_sent'
                self.analysis.save()
            
            return self.success({'message': 'Email başarıyla gönderildi'})
            
        except Exception as e:
            return self.error(f"Email gönderilirken hata: {str(e)}")
    
    # ==========================================
    # Geriye Uyumluluk (Eski metodlar)
    # ==========================================
    
    def process_excel_files(self, source_path: str) -> ServiceResult:
        """
        Belirtilen path'teki Excel dosyalarını işle (geriye uyumluluk)
        """
        if not OPENPYXL_AVAILABLE and not PANDAS_AVAILABLE:
            return self.error("Excel işleme için openpyxl veya pandas kütüphanesi gerekli.")
        
        if not os.path.exists(source_path):
            return self.error(f"Kaynak klasör bulunamadı: {source_path}")
        
        # Excel dosyalarını bul
        excel_patterns = ['*.xlsx', '*.xls', '*.xlsm']
        excel_files = []
        for pattern in excel_patterns:
            excel_files.extend(glob.glob(os.path.join(source_path, pattern)))
        
        if not excel_files:
            return self.error(f"Klasörde Excel dosyası bulunamadı: {source_path}")
        
        processed_count = 0
        total_gids = 0
        total_excel_files = len(excel_files)
        
        for idx, file_path in enumerate(excel_files, 1):
            try:
                gids = self._extract_gids_from_column_d(file_path)
                if gids:
                    self._gids_from_files.update([g.gid for g in gids])
                    self._gid_records.extend(gids)
                    self._processed_files.append(file_path)
                    total_gids += len(gids)
                    processed_count += 1
                    # Progress güncelle
                    progress = int((idx / total_excel_files) * 100)
                    self.update_progress('processing', progress,
                                       f'{idx}/{total_excel_files} dosya işlendi',
                                       {'processed': processed_count, 'total_gids': total_gids})
            except Exception as e:
                self.logger.warning(f"Dosya işlenirken hata: {file_path} - {str(e)}")
                continue
        
        self.update_progress('processing', 100, 'Dosyalar başarıyla işlendi',
                           {'processed': processed_count, 'unique_gids': len(self._gids_from_files)})
        
        return self.success({
            'processed_files': processed_count,
            'total_gids': total_gids,
            'unique_gids': len(self._gids_from_files),
            'files': self._processed_files
        })
    
    def compare_with_system_gids(self) -> ServiceResult:
        """Geriye uyumluluk - compare_with_user_gids'e yönlendir"""
        return self.compare_with_user_gids()
    
    def generate_user_checklist(self) -> ServiceResult:
        """Kullanıcı kontrol listesi Excel dosyası oluştur"""
        if not OPENPYXL_AVAILABLE:
            return self.error("Excel oluşturmak için openpyxl kütüphanesi gerekli.")
        
        if not self._gid_records:
            return self.error("İşlenmiş GID kaydı bulunamadı.")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Kullanıcı Kontrol Listesi"
        
        headers = ['GID', 'Kaynak Dosya', 'Tarih']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row_num, record in enumerate(self._gid_records, 2):
            ws.cell(row=row_num, column=1, value=record.gid)
            ws.cell(row=row_num, column=2, value=record.source_file or '')
            ws.cell(row=row_num, column=3, value=record.date or '')
        
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return self.success({
            'content': output.getvalue(),
            'filename': f"kullanici_kontrol_listesi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        })
    
    def generate_unique_gids_file(self) -> ServiceResult:
        """Unique GID'leri Excel dosyası olarak oluştur"""
        if not OPENPYXL_AVAILABLE:
            return self.error("Excel oluşturmak için openpyxl kütüphanesi gerekli.")
        
        if not self._gids_from_files:
            return self.error("İşlenmiş GID bulunamadı.")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Unique GID'ler"
        
        ws.cell(row=1, column=1, value='GID').font = Font(bold=True)
        
        for row_num, gid in enumerate(sorted(self._gids_from_files), 2):
            ws.cell(row=row_num, column=1, value=gid)
        
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return self.success({
            'content': output.getvalue(),
            'filename': f"unique_gids_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        })
    
    def generate_log_file(self, discrepancies: List[Dict] = None) -> ServiceResult:
        """Farklılıkları log dosyası olarak oluştur"""
        from it_tools.models import GIDDiscrepancy
        
        if discrepancies is None and self.analysis:
            discrepancies = list(GIDDiscrepancy.objects.filter(
                analysis=self.analysis
            ).values('gid', 'discrepancy_type', 'source_file', 'details'))
        
        if not discrepancies:
            return self.error("Kaydedilecek farklılık bulunamadı.")
        
        now = datetime.now()
        filename = f"log-{now.day}_{now.month}_{now.strftime('%y')}.txt"
        
        lines = [
            "AD Log Analiz Raporu",
            f"Tarih: {now.strftime('%d/%m/%Y %H:%M:%S')}",
            "=" * 60,
            "",
            f"Toplam Eşleşmeyen GID Sayısı: {len(discrepancies)}",
            "",
            "-" * 60,
            "EŞLEŞMEYEN GID'LER (AD'de var, Sistemde yok)",
            "-" * 60,
        ]
        
        for d in discrepancies:
            gid = d.get('gid', d) if isinstance(d, dict) else str(d)
            source = d.get('source_file', 'N/A') if isinstance(d, dict) else 'N/A'
            lines.append(f"  - {gid} (Kaynak: {source})")
        
        lines.extend([
            "",
            "=" * 60,
            "Rapor Sonu",
        ])
        
        content = '\n'.join(lines)
        
        return self.success({
            'content': content.encode('utf-8'),
            'filename': filename
        })
